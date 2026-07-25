"""Validation Report Generator — produces professional reports in PDF, Excel, and CSV."""

from __future__ import annotations

import csv
import io
import os
from datetime import datetime, timezone

from validation.engine import ValidationResult, ValidationStatus


class ValidationReportGenerator:
    """Generates validation reports in multiple formats."""

    @staticmethod
    def generate_summary(result: ValidationResult) -> dict:
        """Generate a structured summary dict."""
        findings_by_severity = {"error": [], "warning": [], "info": []}
        for f in result.all_findings:
            sev = f.get("severity", "info")
            findings_by_severity.setdefault(sev, []).append(f)

        findings_by_category = {}
        for f in result.all_findings:
            cat = f.get("category", "unknown")
            findings_by_category.setdefault(cat, []).append(f)

        recommendations = []
        for f in result.all_findings:
            fix = f.get("suggested_fix")
            if fix:
                recommendations.append({
                    "rule": f.get("rule_name"),
                    "severity": f.get("severity"),
                    "fix": fix,
                    "affected_rows": f.get("affected_rows", 0),
                })

        return {
            "dataset_name": result.dataset_name,
            "validation_date": result.validated_at,
            "status": result.status.value,
            "quality_score": result.quality_score.to_dict() if result.quality_score else None,
            "summary": {
                "total_errors": result.total_errors,
                "total_warnings": result.total_warnings,
                "total_info": result.total_info,
                "total_findings": result.total_errors + result.total_warnings + result.total_info,
                "row_count": result.profile.row_count,
                "column_count": result.profile.column_count,
            },
            "findings_by_severity": {
                sev: len(items) for sev, items in findings_by_severity.items()
            },
            "findings_by_category": {
                cat: len(items) for cat, items in findings_by_category.items()
            },
            "recommendations": recommendations,
            "schema_issues": result.schema_result.to_dict(),
            "profile": result.profile.to_dict(),
        }

    @staticmethod
    def export_csv(result: ValidationResult, file_path: str | None = None) -> str:
        """Export findings as CSV. Returns file path or CSV string."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Rule Name", "Category", "Severity", "Column",
            "Affected Rows", "Message", "Suggested Fix", "Business Impact",
        ])
        for f in result.all_findings:
            writer.writerow([
                f.get("rule_name", ""),
                f.get("category", ""),
                f.get("severity", ""),
                f.get("column", ""),
                f.get("affected_rows", 0),
                f.get("message", ""),
                f.get("suggested_fix", ""),
                f.get("business_impact", ""),
            ])

        content = output.getvalue()
        if file_path:
            with open(file_path, "w", newline="", encoding="utf-8") as fh:
                fh.write(content)
            return file_path
        return content

    @staticmethod
    def export_excel(result: ValidationResult, file_path: str) -> str:
        """Export report as Excel file with multiple sheets."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        summary = ValidationReportGenerator.generate_summary(result)

        ws_summary["A1"] = "Hospital Data Validation Report"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A3"] = "Dataset"
        ws_summary["B3"] = summary["dataset_name"]
        ws_summary["A4"] = "Date"
        ws_summary["B4"] = summary["validation_date"]
        ws_summary["A5"] = "Status"
        ws_summary["B5"] = summary["status"]
        ws_summary["A6"] = "Quality Score"
        if summary["quality_score"]:
            ws_summary["B6"] = summary["quality_score"]["overall"]
            ws_summary["C6"] = summary["quality_score"]["traffic_light"]
        ws_summary["A8"] = "Errors"
        ws_summary["B8"] = summary["summary"]["total_errors"]
        ws_summary["A9"] = "Warnings"
        ws_summary["B9"] = summary["summary"]["total_warnings"]
        ws_summary["A10"] = "Info"
        ws_summary["B10"] = summary["summary"]["total_info"]
        ws_summary["A12"] = "Rows"
        ws_summary["B12"] = summary["summary"]["row_count"]
        ws_summary["A13"] = "Columns"
        ws_summary["B13"] = summary["summary"]["column_count"]

        # Color code status
        status_colors = {
            "passed": "00FF00",
            "passed_with_warnings": "FFFF00",
            "failed": "FF0000",
            "approved": "00FF00",
            "rejected": "FF0000",
        }
        fill = PatternFill(start_color=status_colors.get(summary["status"], "FFFFFF"), end_color=status_colors.get(summary["status"], "FFFFFF"), fill_type="solid")
        ws_summary["B5"].fill = fill

        # Findings sheet
        ws_findings = wb.create_sheet("Findings")
        headers = ["Rule Name", "Category", "Severity", "Column", "Affected Rows", "Message", "Suggested Fix", "Business Impact"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_findings.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        for row_idx, f in enumerate(result.all_findings, 2):
            ws_findings.cell(row=row_idx, column=1, value=f.get("rule_name", ""))
            ws_findings.cell(row=row_idx, column=2, value=f.get("category", ""))
            ws_findings.cell(row=row_idx, column=3, value=f.get("severity", ""))
            ws_findings.cell(row=row_idx, column=4, value=f.get("column", ""))
            ws_findings.cell(row=row_idx, column=5, value=f.get("affected_rows", 0))
            ws_findings.cell(row=row_idx, column=6, value=f.get("message", ""))
            ws_findings.cell(row=row_idx, column=7, value=f.get("suggested_fix", ""))
            ws_findings.cell(row=row_idx, column=8, value=f.get("business_impact", ""))

        # Quality Score sheet
        if summary["quality_score"]:
            ws_score = wb.create_sheet("Quality Score")
            ws_score["A1"] = "Dimension"
            ws_score["B1"] = "Score"
            ws_score["A1"].font = Font(bold=True)
            ws_score["B1"].font = Font(bold=True)
            for row_idx, (key, val) in enumerate(summary["quality_score"].items(), 2):
                ws_score.cell(row=row_idx, column=1, value=key)
                ws_score.cell(row=row_idx, column=2, value=val)

        # Profile sheet
        ws_profile = wb.create_sheet("Data Profile")
        ws_profile["A1"] = "Metric"
        ws_profile["B1"] = "Value"
        ws_profile["A1"].font = Font(bold=True)
        ws_profile["B1"].font = Font(bold=True)
        ws_profile["A2"] = "Row Count"
        ws_profile["B2"] = summary["summary"]["row_count"]
        ws_profile["A3"] = "Column Count"
        ws_profile["B3"] = summary["summary"]["column_count"]
        ws_profile["A4"] = "Overall Completeness"
        ws_profile["B4"] = f"{summary['profile']['overall_completeness']:.1f}%"
        ws_profile["A5"] = "Overall Uniqueness"
        ws_profile["B5"] = f"{summary['profile']['overall_uniqueness']:.1f}%"
        ws_profile["A6"] = "Duplicate Percentage"
        ws_profile["B6"] = f"{summary['profile']['duplicate_percentage']:.1f}%"

        col_start = 8
        ws_profile.cell(row=col_start, column=1, value="Column").font = Font(bold=True)
        ws_profile.cell(row=col_start, column=2, value="Dtype").font = Font(bold=True)
        ws_profile.cell(row=col_start, column=3, value="Null %").font = Font(bold=True)
        ws_profile.cell(row=col_start, column=4, value="Unique Count").font = Font(bold=True)
        for row_idx, col_data in enumerate(summary["profile"]["columns"], col_start + 1):
            ws_profile.cell(row=row_idx, column=1, value=col_data["name"])
            ws_profile.cell(row=row_idx, column=2, value=col_data["dtype"])
            ws_profile.cell(row=row_idx, column=3, value=f"{col_data['null_percentage']:.1f}%")
            ws_profile.cell(row=row_idx, column=4, value=col_data["unique_count"])

        wb.save(file_path)
        return file_path

    @staticmethod
    def export_pdf(result: ValidationResult, file_path: str) -> str:
        """Export report as PDF."""
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Hospital Data Validation Report", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(5)

        pdf.set_font("Helvetica", size=10)
        summary = ValidationReportGenerator.generate_summary(result)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)

        pdf.cell(60, 6, "Dataset:")
        pdf.cell(0, 6, summary["dataset_name"], new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Validation Date:")
        pdf.cell(0, 6, summary["validation_date"], new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Status:")
        pdf.cell(0, 6, summary["status"], new_x="LMARGIN", new_y="NEXT")
        if summary["quality_score"]:
            pdf.cell(60, 6, "Quality Score:")
            pdf.cell(0, 6, f"{summary['quality_score']['overall']:.1f} ({summary['quality_score']['traffic_light']})", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Rows:")
        pdf.cell(0, 6, str(summary["summary"]["row_count"]), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Columns:")
        pdf.cell(0, 6, str(summary["summary"]["column_count"]), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Findings Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(60, 6, "Total Errors:")
        pdf.cell(0, 6, str(summary["summary"]["total_errors"]), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Total Warnings:")
        pdf.cell(0, 6, str(summary["summary"]["total_warnings"]), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(60, 6, "Total Info:")
        pdf.cell(0, 6, str(summary["summary"]["total_info"]), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # Quality scores
        if summary["quality_score"]:
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Quality Scores", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", size=10)
            for key, val in summary["quality_score"].items():
                pdf.cell(60, 6, f"{key.replace('_', ' ').title()}:")
                pdf.cell(0, 6, str(val), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

        # Findings detail
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Findings Detail", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=8)

        for f in result.all_findings[:50]:  # Limit to first 50
            sev = f.get("severity", "info").upper()
            msg = f.get("message", "")
            pdf.set_text_color(255 if sev == "ERROR" else 0, 0, 0)
            pdf.multi_cell(0, 5, f"[{sev}] {f.get('rule_name', '')}: {msg}")
            if f.get("suggested_fix"):
                pdf.set_text_color(0, 0, 200)
                pdf.multi_cell(0, 5, f"  Fix: {f['suggested_fix']}")
            pdf.set_text_color(0, 0, 0)

        if len(result.all_findings) > 50:
            pdf.ln(3)
            pdf.cell(0, 6, f"... and {len(result.all_findings) - 50} more findings (see Excel/CSV for full report).", new_x="LMARGIN", new_y="NEXT")

        pdf.output(file_path)
        return file_path
